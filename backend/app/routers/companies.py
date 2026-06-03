import csv
import io
from typing import List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from ..models.company import Company
from ..models.founder import Founder
from ..models.signal import Signal, SignalType
from ..models.user import User
from ..routers.auth import get_current_user
from ..schemas.company import CompanyCreate, CompanyResponse, CompanyUpdate
from ..schemas.signal import SignalResponse

router = APIRouter(prefix="/companies", tags=["companies"])


def _company_response(company: Company, db: Session) -> CompanyResponse:
    signal_count = db.query(func.count(Signal.id)).filter(Signal.company_id == company.id).scalar()
    resp = CompanyResponse.model_validate(company)
    resp.signal_count = signal_count
    return resp


def _signal_response(signal: Signal) -> SignalResponse:
    resp = SignalResponse.model_validate(signal)
    if signal.company:
        resp.company_name = signal.company.name
    return resp


@router.get("", response_model=List[CompanyResponse])
def list_companies(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(Company)
    if active_only:
        q = q.filter(Company.is_active)
    companies = q.order_by(Company.name).all()
    return [_company_response(c, db) for c in companies]


@router.post("", response_model=CompanyResponse, status_code=status.HTTP_201_CREATED)
def create_company(
    company_in: CompanyCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    founder_list = company_in.founders or []
    company_data = company_in.model_dump(exclude={'founders'})
    company = Company(**company_data)
    db.add(company)
    db.flush()
    for f in founder_list:
        if f.name.strip():
            db.add(Founder(company_id=company.id, **f.model_dump()))
    db.commit()
    db.refresh(company)
    return _company_response(company, db)


@router.get("/import-template")
def download_import_template(
    current_user: User = Depends(get_current_user),
):
    """Download a pre-filled Excel template for bulk company import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    headers = ["name", "website", "linkedin_url", "category", "description"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 28 if header in ("name", "description") else 22

    # Add a note on the category cell explaining valid values
    from openpyxl.comments import Comment
    cat_cell = ws.cell(row=1, column=4)
    cat_cell.comment = Comment(
        "Valid values (one per row):\nFund 1 Portfolio\nFund 2 Portfolio\nFund 3 Portfolio\nUnicorn\nKeep Close",
        "Template",
    )

    ws.append(["Acme Corp", "https://acme.com", "https://linkedin.com/company/acme",
               "Fund 1 Portfolio", "B2B workflow automation platform"])
    ws.append(["Beta AI", "https://betaai.io", "", "Unicorn", "LLM-based document processing"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=company_import_template.xlsx"},
    )


@router.post("/import")
async def import_companies(
    file: UploadFile,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Bulk import companies from CSV or XLSX. Admin only."""
    if current_user.role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    filename = file.filename or ""
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are accepted")

    content = await file.read()
    rows: list[dict] = []

    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        header_row = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({header_row[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower(): (v.strip() if v else "") for k, v in row.items()})

    imported, skipped, errors = 0, 0, []

    for i, row in enumerate(rows, start=2):
        name = row.get("name", "").strip()
        if not name:
            errors.append(f"Row {i}: missing required field 'name'")
            continue

        website = row.get("website", "").strip() or None

        dup_filter = [func.lower(Company.name) == name.lower()]
        if website:
            dup_filter.append(func.lower(Company.website) == website.lower())
        existing = db.query(Company).filter(or_(*dup_filter)).first()
        if existing:
            skipped += 1
            continue

        raw_category = row.get("category", "").strip()
        valid_categories = {"Fund 1 Portfolio", "Fund 2 Portfolio", "Fund 3 Portfolio", "Unicorn", "Keep Close"}
        categories = [raw_category] if raw_category in valid_categories else []

        company = Company(
            name=name,
            website=website,
            linkedin_url=row.get("linkedin_url", "").strip() or None,
            categories=categories or None,
            description=row.get("description", "").strip() or None,
        )
        db.add(company)
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.get("/{company_id}", response_model=CompanyResponse)
def get_company(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")
    return _company_response(company, db)


@router.put("/{company_id}", response_model=CompanyResponse)
def update_company(
    company_id: int,
    company_in: CompanyUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")

    update_data = company_in.model_dump(exclude_unset=True)
    founder_list = update_data.pop('founders', None)

    for field, value in update_data.items():
        setattr(company, field, value)

    if founder_list is not None:
        existing = {
            f.name.lower(): f
            for f in db.query(Founder).filter(Founder.company_id == company_id).all()
        }
        submitted = set()
        for fd in founder_list:
            fd_dict = fd if isinstance(fd, dict) else fd.model_dump()
            name = fd_dict.get('name', '').strip()
            if not name:
                continue
            name_lower = name.lower()
            submitted.add(name_lower)
            if name_lower in existing:
                for k, v in fd_dict.items():
                    setattr(existing[name_lower], k, v)
            else:
                db.add(Founder(company_id=company_id, **fd_dict))
        for name_lower, founder in existing.items():
            if name_lower not in submitted:
                db.delete(founder)

    db.commit()
    db.refresh(company)
    return _company_response(company, db)


@router.delete("/{company_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_company(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")
    db.delete(company)
    db.commit()


@router.get("/{company_id}/signals", response_model=List[SignalResponse])
def get_company_signals(
    company_id: int,
    signal_type: Optional[SignalType] = None,
    limit: int = Query(100, le=500),
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")

    q = db.query(Signal).options(joinedload(Signal.company)).filter(Signal.company_id == company_id)
    if signal_type:
        q = q.filter(Signal.signal_type == signal_type)

    signals = q.order_by(Signal.detected_at.desc()).offset(offset).limit(limit).all()
    return [_signal_response(s) for s in signals]


@router.post("/{company_id}/refresh", status_code=status.HTTP_202_ACCEPTED)
async def refresh_company(
    company_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")

    from ..services.scheduler import process_single_company
    background_tasks.add_task(process_single_company, company_id)
    return {"message": f"Refresh queued for {company.name}"}
